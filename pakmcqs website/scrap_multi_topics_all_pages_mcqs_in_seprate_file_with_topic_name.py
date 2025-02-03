import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import re
from civil_eng_multi_topics import multi_topics  # Importing the topics list

def sanitize_filename(filename):
    """Replace invalid characters in filenames with underscores."""
    return re.sub(r'[\/:*?"<>|]', '_', filename)

def scrape_mcqs(url, start_idx, topic_title):
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')
    if soup.title.text == "Page not found - PakMcqs":
        print("Page title: ", soup.title.text)
        return None, start_idx

    all_mcqs = soup.find_all("article", class_="l-post grid-post grid-card-post")
    mcq_list = []
    for mcq in all_mcqs:
        item = {}
        item["#Sr."] = start_idx
        item["topic"] = topic_title
        item["question"] = mcq.find('h2').text.strip()
        
        options_container = mcq.find("div", class_="excerpt")
        options_html = options_container.find('p')
        options = options_html.get_text(separator="\n").strip().split('\n')
        options = [opt.strip() for opt in options if opt.strip()]  # Remove empty options and strip extra spaces
        options = [opt[3:].strip() if len(opt) > 1 and opt[1] == '.' else opt for opt in options]  # Remove option labels

        # Ignore the options if the correct one is not found and place the cell empty
        correct_option_text = ""
        if options_html.find('strong') is not None:
            correct_option_text = options_html.find('strong').text.strip()
            correct_option_text = correct_option_text[3:].strip() if len(correct_option_text) > 1 and correct_option_text[1] == '.' else correct_option_text  # Remove option labels from correct option

        # Assign options to respective columns
        item["A"] = options[0] if len(options) > 0 else ""
        item["B"] = options[1] if len(options) > 1 else ""
        item["C"] = options[2] if len(options) > 2 else ""
        item["D"] = options[3] if len(options) > 3 else ""

        # Determine the correct option letter
        if correct_option_text == item["A"]:
            item["Ans"] = "A"
        elif correct_option_text == item["B"]:
            item["Ans"] = "B"
        elif correct_option_text == item["C"]:
            item["Ans"] = "C"
        elif correct_option_text == item["D"]:
            item["Ans"] = "D"
        else:
            item["Ans"] = ""

        if item["Ans"]:  # Only save MCQs that have an answer
            mcq_list.append(item)
        start_idx += 1
    
    return mcq_list, start_idx

# Define the folder path where files will be saved
output_folder = "pak_study_mcqs"
os.makedirs(output_folder, exist_ok=True)  # Create the folder if it does not exist

start_idx = 1
headers = ["#Sr.", "Question", "A", "B", "C", "D", "Ans"]
header_alignment = Alignment(horizontal='center')

for topic in multi_topics:
    print(f"=> {topic['id']} Extracting the topic: {topic['title']}")
    current_page = 1
    proceed = True
    topic_data = []
    
    while proceed:
        print(f"\t\t - Scraping page {current_page} of topic '{topic['title']}'")
        url = topic['url'].replace('page/1', f'page/{current_page}')
        print(f"\t\t\t - Extracting URL: {url}")
        
        mcqs_data, start_idx = scrape_mcqs(url, start_idx, topic['title'])
        if mcqs_data is None:
            proceed = False
        else:
            print(f"\t\t\t - Total extracted MCQs from page {current_page}: {len(mcqs_data)}")
            topic_data.extend(mcqs_data)
            current_page += 1
        
        print(f"\t\t\t - Overall MCQs for topic '{topic['title']}': {len(topic_data)}")

    if topic_data:
        # Create a new workbook for each topic
        wb = Workbook()
        ws = wb.active
        ws.title = "MCQs"

        # Add topic title
        ws.append([topic['title']])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
        ws.cell(row=ws.max_row, column=1).alignment = header_alignment

        # Append the headers
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.alignment = header_alignment

        # Append data rows
        for item in topic_data:
            ws.append([item["#Sr."], item["question"], item["A"], item["B"], item["C"], item["D"], item["Ans"]])

        # Save the file inside the subfolder with a sanitized topic name
        safe_topic_name = sanitize_filename(topic['title'])
        file_path = os.path.join(output_folder, f"{safe_topic_name}.xlsx")
        wb.save(file_path)
        print(f"Saved {file_path}")

print("Scraping complete. All topics saved in the 'pak_study_mcqs' folder.")
